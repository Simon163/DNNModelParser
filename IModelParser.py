import numpy as np


class IModelParser():
    '''
    interface of model parser
    '''

    def __init__(self):
        self.ops_info = []
        self.max_input_length = 0
        self.max_output_length = 0

    def _is_conv_or_convdw(self, op_info):
        '''
        check if the op is convolution layer or depthwise convolution layer,
        if yes, return True, else return False
        '''
        if self._is_conv(op_info) or self._is_convdw(op_info):
            return True
        else:
            return False

    def _is_conv(self, op_info):
        '''
        check if the op is convolution layer,
        if yes, return True, else return False
        '''
        if op_info['Op_Type'] in [
                "Convolution",  # caffe
                "Conv2D",  # tf
                "Conv",  #onnx
                "ConvTranspose",  # onnx
        ]:
            return True
        else:
            return False

    def _is_convdw(self, op_info):
        '''
        check if the op is depthwise convolution layer,
        if yes, return True, else return False
        '''
        if op_info['Op_Type'] in [
                "ConvolutionDepthwise",  # caffe
                "DepthwiseConv2dNative",  # tf
        ]:
            return True
        else:
            return False

    def _get_sum_flops(self):
        '''
        get the sum of FLOPs of all ops in the model
        '''
        sum_flops = 0
        for op_info in self.ops_info:
            sum_flops += op_info["FLOPs"]
        return sum_flops

    def _get_empty_op_info(self):
        return {
            "Op_ID": "",
            "Op_Type": "",
            "Op_Name": "",
            "FLOPs": "",
            "Input_Dimension_X": [],
            "Input_Dimension_Y": [],
            "Input_Dimension_C": [],
            "Output_Dimension_X": [],
            "Output_Dimension_Y": [],
            "Output_Dimension_C": [],
            "Kernel_Dimension_X": "",
            "Kernel_Dimension_Y": "",
            'Kernel_Outliers': "",
            "Pad_Dimension_X": "",
            "Pad_Dimension_Y": "",
            "Stride_Dimension_X": "",
            "Stride_Dimension_Y": "",
            "Dilate_Dimension_X": "",
            "Dilate_Dimension_Y": "",
            "Pad_Mode": "",
            "Group": "",
            "Pooling_Mode": "",
            "Activate_Mode": ""
        }

    def _get_op_info(self):
        pass

    def _check_outliers(self, array):
        ''' check if array has nan or inf, if so, set to 0,
            return the number of nan and inf, and max, min, mean of array
            Args:
                array: numpy array
            Returns:
                list: the number of nan and inf, and max, min, mean of array
        '''
        # check nan
        where_nan = np.where(np.isnan(array))
        num_nan = len(where_nan[0])
        for k in range(len(where_nan[0])):
            array[[index[k] for index in where_nan]] = 0

        #check inf
        where_inf = np.where(np.isinf(array))
        num_inf = len(where_inf[0])
        for k in range(len(where_inf[0])):
            array[[index[k] for index in where_inf]] = 0

        mean = array.mean()
        max = array.max()
        min = array.min()
        return [num_nan, num_inf, max, min, mean]

    def _write_excel(self, excel_name, work_book=None, sheet_name=None):
        '''
        write the model information to an excel file
        '''

        def change_column_dimensions(ws):
            ws.column_dimensions['A'].width = 11.0  # Op ID
            ws.column_dimensions['B'].width = 21.0  # Op Type
            ws.column_dimensions['D'].width = 16.0  # Kernel Shape
            for i in range(self.max_input_length):
                ws.column_dimensions[chr(ord('E') +
                                         i)].width = 16.0  # Input Shape
            column_start = chr(ord('E') + self.max_input_length)
            for i in range(self.max_output_length):
                ws.column_dimensions[chr(ord(column_start) +
                                         i)].width = 16.0  # Output Shape
            column_start = chr(ord(column_start) + self.max_output_length)
            ws.column_dimensions[chr(ord(column_start) +
                                     6)].width = 11.0  # Kernel Mem
            ws.column_dimensions[chr(ord(column_start) +
                                     7)].width = 11.0  # Output Mem
            ws.column_dimensions[chr(ord(column_start) +
                                     8)].width = 11.0  # FLOPs rate
            ws.column_dimensions[chr(ord(column_start) +
                                     9)].width = 11.0  # Kernel nan
            ws.column_dimensions[chr(ord(column_start) +
                                     10)].width = 11.0  # kernel inf
            ws.column_dimensions[chr(ord(column_start) +
                                     11)].width = 11.0  # kerner max
            ws.column_dimensions[chr(ord(column_start) +
                                     12)].width = 11.0  # kerner min
            ws.column_dimensions[chr(ord(column_start) +
                                     13)].width = 11.0  # kerner mean

        def write_cell(ws, row_start, column_start, value, dir):
            for i in range(len(value)):
                if dir == 'column':
                    ws.cell(row=row_start,
                            column=column_start + i,
                            value=value[i])
                elif dir == 'row':
                    ws.cell(row=row_start + i,
                            column=column_start,
                            value=value[i])
                else:
                    print('error: write to cell, failed')

        def get_part_kernel_shape(op_info):
            if self._is_conv_or_convdw(op_info):
                return [
                    "({},{},{},{})".format(
                        op_info["Kernel_Dimension_X"],
                        op_info["Kernel_Dimension_Y"],
                        op_info["Input_Dimension_C"][0]
                        if op_info["Input_Dimension_C"] else ' ',
                        op_info["Output_Dimension_C"][0]
                        if op_info["Output_Dimension_C"] else ' ')
                ]
            elif op_info["Op_Type"] in [
                    "AvgPool",
                    "Pooling",  #caffe
                    "MaxPool",
                    "AveragePool"  #onnx
            ]:
                return [
                    "({},{})".format(op_info["Kernel_Dimension_X"],
                                     op_info["Kernel_Dimension_Y"])
                ]
            else:
                return [""]

        def get_part_padding(op_info):
            if op_info["Pad_Dimension_X"]:
                return [
                    "({},{})".format(op_info["Pad_Dimension_X"],
                                     op_info["Pad_Dimension_Y"])
                ]
            else:
                return [""]

        def get_part_inoutput_shape(op_info, attr="input"):
            part3 = []
            if attr == "input":
                tensor_num = len(op_info["Input_Dimension_X"])
                for i in range(tensor_num):
                    part3 += [
                        "({},{},{})".format(
                            op_info["Input_Dimension_X"][i],
                            op_info["Input_Dimension_Y"][i]
                            if op_info["Input_Dimension_Y"] else "",
                            op_info["Input_Dimension_C"][i]
                            if op_info["Input_Dimension_C"] else "")
                    ]
                if tensor_num < self.max_input_length:
                    part3 += (self.max_input_length - tensor_num) * [""]
            elif attr == "output":
                tensor_num = len(op_info["Output_Dimension_X"])
                for i in range(tensor_num):
                    part3 += [
                        "({},{},{})".format(
                            op_info["Output_Dimension_X"][i],
                            op_info["Output_Dimension_Y"][i]
                            if op_info["Output_Dimension_Y"] else "",
                            op_info["Output_Dimension_C"][i]
                            if op_info["Output_Dimension_C"] else "")
                    ]
                if tensor_num < self.max_input_length:
                    part3 += (self.max_output_length - tensor_num) * [""]
            return part3

        from openpyxl import Workbook

        # create an empty Workbook and add 'readme' info.
        wb = Workbook() if work_book is None else work_book
        ws = wb.active
        ws.title = "readme"
        write_cell(ws, 1, 1, ["ONNX model parser"], 'row')
        write_cell(ws, 2, 1, ["Original: before graph optimization"], 'row')
        write_cell(ws, 3, 1, ["vs."], 'row')
        write_cell(ws, 4, 1, ["Optimized: after graph optimization"], 'row')

        # add a worksheet.
        ws = wb.active if sheet_name is None else wb.create_sheet(sheet_name)
        change_column_dimensions(ws)

        # ========network structure analysis========
        network_structure_title = ["Op ID", "Op Type", "Op Name", "Kernel Shape"] + \
                                  self.max_input_length * ["Input Shape"] + \
                                  self.max_output_length * ["Output Shape"] + \
                                  ["Padding", "Stride", "Dilation", "Pad Mode", "Pooling", "Activate"]
        ws.cell(row=1, column=1, value=">>>> Network Structure")
        row_start = 2
        column_start = 1
        write_cell(ws, row_start, column_start, network_structure_title,
                   'column')

        for index, op_info in enumerate(self.ops_info):
            part1 = [op_info['Op_ID'], op_info['Op_Type'], op_info["Op_Name"]]
            part2 = get_part_kernel_shape(op_info)
            part3 = get_part_inoutput_shape(op_info, "input")
            part4 = get_part_inoutput_shape(op_info, "output")
            part5 = get_part_padding(op_info)
            part6 = [
                "({},{})".format(op_info["Stride_Dimension_X"],
                                 op_info["Stride_Dimension_Y"])
            ]
            part7 = [
                "({},{})".format(op_info["Dilate_Dimension_X"],
                                 op_info["Dilate_Dimension_Y"])
            ]
            part8 = [
                op_info["Pad_Mode"], op_info["Pooling_Mode"],
                op_info["Activate_Mode"]
            ]
            part = part1
            part += part2
            part += part3 + part4
            part += part5
            part += part6 if (self._is_conv_or_convdw(op_info)
                              or op_info["Op_Type"] == "AvgPool") else [""]
            part += part7 if (self._is_conv_or_convdw(op_info)) else [""]
            part += part8

            row_start = index + 3
            column_start = 1
            write_cell(ws, row_start, column_start, part, 'column')

        # ========network memory analysis========
        row_start = 3
        col_start = len(network_structure_title) + 1
        ws.cell(row=1, column=col_start, value=">>>> Network Memory Analysis")
        Network_Analysis_Title = [
            "Kernel Mem", "Output Mem", "FLOPs", "FLOPs Rate"
        ]
        write_cell(ws, 2, col_start, Network_Analysis_Title, 'column')

        element_bytes = 4  # float, 4 bytes
        sum_kernel_memory = 0  # bytes
        sum_output_memory = 0  # bytes
        sum_flops = self._get_sum_flops()
        for index, op_info in enumerate(self.ops_info):
            part1 = []
            if self._is_conv(op_info):
                tmp_kernel_memory = op_info["Kernel_Dimension_X"] * op_info["Kernel_Dimension_Y"] * \
                                    op_info["Output_Dimension_C"][0] * op_info["Input_Dimension_C"][0]
                tmp_kernel_memory *= element_bytes
                part1 += [tmp_kernel_memory]
            elif self._is_convdw(op_info):
                tmp_kernel_memory = op_info["Kernel_Dimension_X"] * op_info["Kernel_Dimension_Y"] * \
                                    op_info["Input_Dimension_C"][0]
                tmp_kernel_memory *= element_bytes
                part1 += [tmp_kernel_memory]
            else:
                tmp_kernel_memory = 0
                part1 += ['']
            sum_kernel_memory += tmp_kernel_memory
            part2 = []
            tmp_output_memory = 0
            for i in range(len(op_info["Output_Dimension_X"])):
                tmp_output_memory += op_info["Output_Dimension_X"][i] * (op_info["Output_Dimension_Y"][i] if op_info["Output_Dimension_Y"] else 1) * \
                                     (op_info["Output_Dimension_C"][i] if op_info["Output_Dimension_C"] else 1)
            tmp_output_memory *= element_bytes
            sum_output_memory += tmp_output_memory
            part2 += [tmp_output_memory]
            part3 = [
                round(op_info["FLOPs"] * 1024 * 1024, 0),
                round(op_info["FLOPs"] / sum_flops *
                      100, 2) if sum_flops else 0
            ]
            part = part1 + part2 + part3
            write_cell(ws, index + row_start, col_start, part, 'column')

        # ========network outliers========
        row_start = 3
        col_start = len(network_structure_title) + len(
            Network_Analysis_Title) + 1
        ws.cell(row=1, column=col_start, value=">>>> Network Outliers")
        Network_Outliers_Title = [
            "Kernel Nan", "Kernel Inf", "Kernel Max", "Kernel Min",
            "Kernel Mean", "Bias Nan", "Bias Inf", "Bias Max", "Bias Min",
            "Bias Mean"
        ]
        write_cell(ws, 2, col_start, Network_Outliers_Title, 'column')
        for index, op_info in enumerate(self.ops_info):
            part1 = []
            if op_info["Kernel_Outliers"]:
                part1 += [i for i in op_info["Kernel_Outliers"]]
            part = part1
            write_cell(ws, index + row_start, col_start, part, 'column')

        # summary
        row_start = 2 + len(self.ops_info) + 1 + 1
        ws.cell(row=row_start, column=1, value=">>>> Network Summary")
        row_start += 1
        Network_Summary_Title = ["Kernel Mem", "Output Mem", "FLOPs"]
        for i in range(len(Network_Summary_Title)):
            ws.cell(row=row_start + i,
                    column=1,
                    value=Network_Summary_Title[i])
        part1 = [
            str(round(sum_kernel_memory / (1024 * 1024), 2)) + "(M)",
            str(round(sum_output_memory / (1024 * 1024), 2)) + "(M)",
            str(round(sum_flops, 2)) + "(M)"
        ]
        part = part1
        write_cell(ws, row_start, 2, part, 'row')

        # save to excel
        wb.save(excel_name)
